import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
wb=load_workbook('grades.xlsx')
ws=wb.active
print(ws)
ws[A2].value='whhwh'
wb.save('grades.xslx')
print([wb.sheetnames])
wb.create_sheet('test')
wb=Workbook()
ws=wb.active
ws.title=('')
ws.append(['any','list'])
ws.merge_cells(['A1:D1'])
ws.insert_rows(7)#inserts rows after 7th row
ws.delete_rows(7)#deletes 7th row
ws.insert_cols(2)#inserts cols at 2nd col
ws.delete_rows(2)#deletes
ws.move_range('C1:D11',rows=2,cols=3)
'''email_sender="avnijain934@gmail.com"
email_pass='taxqkihcpsdccjnv'
email_rec='avni.jain2022@vitstudent.ac.in'
subject="check"
body="""
Okay, lets try to code for real now
jdcjhc
"""
em=EmailMessage()
em["From"]=email_sender
em["To"]=email_rec
em["Subject"]=subject
em.set_content(body)
context=ssl.create_default_context()
with mtplib.SSMTP_SSL('smtp.gmail.com',465,context=context) as smtp:
    smtp.login(email_sender,email_pass)
    smtp.sendmail(email_sender,email_rec,em.as_string())'''
