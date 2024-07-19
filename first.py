import openpyxl
import plyer
from openpyxl import Workbook, load_workbook
import datetime
from datetime import datetime
from openpyxl.utils import get_column_letter
from plyer import notification
import pywhatkit

wb=load_workbook('Resource allocation.xlsx')
ws=wb.active

dt=datetime.now()
d=dt.strftime("%d/%m/%y")
t=dt.strftime("%H:%M:%S")
H=int(dt.strftime("%H"))
M=int(dt.strftime("%M"))
for row in range (2,11):
    ped=ws['E'+str(row)].value
    if ped==None:
        continue
    pedd=ped.strftime("%d/%m/%y")
    if d>pedd:
        if ws['F'+str(row)].value!='Completed':
            M+=1
            pywhatkit.sendwhatmsg('+91'+str(ws['H'+str(row)].value), 'Your task '+ws['C'+str(row)].value+ ' is Overdue',
                              H,M,10)

wb.save("Resource allocation.xlsx")
